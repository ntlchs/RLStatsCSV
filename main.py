import os
import glob
from openpyxl import Workbook, load_workbook
from datetime import datetime

sourceFiles = glob.glob('/mnt/c/Users/narp/AppData/Roaming/bakkesmod/bakkesmod/data/RocketStats/RocketStats_*.txt')

# set first row as stats names and first column as the current date
fileNames = ['CurrentDate'] 
now = datetime.today().strftime('%d-%m-%Y %H:%M')
fileValues = [now]

# collect data

for f in sourceFiles:
    sourceNames = f.replace('/mnt/c/Users/narp/AppData/Roaming/bakkesmod/bakkesmod/data/RocketStats/RocketStats_','')
    sourceNames = sourceNames.replace('.txt','')
    fileNames.append(sourceNames)
    with open(f) as ff:
        sourceValues = ff.readline()
        try:
            fileValues.append(float(sourceValues))
        except ValueError:
            fileValues.append(sourceValues)

# CSV

firstRow = ','.join(fileNames)
nextRow = ','.join([str(v) for v in fileValues])
csvFile = '/home/nat/rl_stats/stats.csv'
if os.path.isfile(csvFile):
    print('csv file exists, stats appended') 
    with open(csvFile,'a') as csv:
        csv.write('\n')
        csv.write(nextRow)        
else:
    print('csv file does not exist, file created')
    with open(csvFile,'w') as csv:
        csv.write(firstRow)
        csv.write('\n')
        csv.write(nextRow)

# XLSX

xlsxFile = '/home/nat/rl_stats/stats.xlsx'
if os.path.isfile(xlsxFile):
    print('xlsx file exists, stats added')
    workbook = load_workbook(filename = xlsxFile)
    worksheet = workbook.active 
    row = 2
    col = 1
    cellValue = worksheet.cell(row = row,column= col).value
    while cellValue is not None:
        row += 1 
        cellValue = worksheet.cell(row = row,column= col).value
    for v in fileValues:
        worksheet.cell(column=col, row=row, value=v)
        col += 1
    workbook.save(xlsxFile)

else:
    print('xlsx file does not exist, new file created')   
    workbook = Workbook()
    worksheet = workbook.active  
    row = 1
    col = 1
    for n in fileNames:
        worksheet.cell(column=col, row=row, value=n)
        col += 1
    row = 2
    col = 1
    for v in fileValues:
        worksheet.cell(column=col, row=row, value=v)
        col += 1
    workbook.save(xlsxFile)